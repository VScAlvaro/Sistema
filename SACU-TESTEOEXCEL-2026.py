# ==========================================
# SIMULACIÓN REALISTA PROCESO SENESCYT
# CARGA DESDE EXCEL CON OPENPYXL
# ==========================================

from openpyxl import load_workbook

# -------------------------------
# CONFIGURACIÓN GENERAL
# -------------------------------
PUNTAJE_MINIMO_GENERAL = 650

SEGMENTOS = {
    "vulnerabilidad": 20,
    "merito": 20,
    "pueblos": 10,
    "bachiller_actual": 30,
    "procesos_anteriores": 20,
    "excluidos": 10,
}

BONIFICACION_MAX = 50


# ===============================
# ASPIRANTE
# ===============================
class Aspirante:
    def __init__(self, cedula, nombre, puntaje, segmentos, opciones):
        self.cedula = cedula
        self.nombre = nombre
        self.puntaje = puntaje
        self.segmentos = segmentos
        self.opciones = opciones
        self.carrera = None
        self.estado = "PENDIENTE"

    def bonificacion(self):
        return min(sum(SEGMENTOS[s] for s in self.segmentos), BONIFICACION_MAX)

    def puntaje_final(self):
        return self.puntaje + self.bonificacion()

    def mostrar(self):
        print(
            f"{self.nombre} | Base: {self.puntaje} | "
            f"Bonif: {self.bonificacion()} | "
            f"Final: {self.puntaje_final()} | "
            f"Asignado: {self.carrera or '-'} | "
            f"Estado: {self.estado}"
        )


# ===============================
# CARRERA
# ===============================
class Carrera:
    def __init__(self, nombre, cupos, puntaje_minimo):
        self.nombre = nombre
        self.cupos = cupos
        self.puntaje_minimo = puntaje_minimo

    def mostrar(self):
        print(
            f"{self.nombre} | Cupos: {self.cupos} | "
            f"Puntaje mínimo: {self.puntaje_minimo}"
        )


# ===============================
# CARGAR ASPIRANTES (OPENPYXL)
# ===============================
def cargar_aspirantes_excel(ruta_excel):
    wb = load_workbook(ruta_excel)
    ws = wb.active

    aspirantes = []

    # Se asume que la fila 1 es encabezado
    for fila in ws.iter_rows(min_row=2, values_only=True):
        (
            identificacion,
            nombres,
            apellidos,
            nota,
            carrera,
            vulnerabilidad,
            merito,
            pueblos,
            bachiller
        ) = fila

        segmentos = []

        if vulnerabilidad == "SI":
            segmentos.append("vulnerabilidad")
        if merito == "SI":
            segmentos.append("merito")
        if pueblos == "SI":
            segmentos.append("pueblos")
        if bachiller == "SI":
            segmentos.append("bachiller_actual")

        aspirantes.append(
            Aspirante(
                cedula=str(identificacion),
                nombre=f"{nombres} {apellidos}",
                puntaje=float(nota),
                segmentos=segmentos,
                opciones=[str(carrera).strip().upper()]
            )
        )

    return aspirantes


# ===============================
# CARGAR CARRERAS (OPENPYXL)
# ===============================
def cargar_carreras_excel(ruta_excel):
    wb = load_workbook(ruta_excel)
    ws = wb.active

    carreras = []

    for fila in ws.iter_rows(min_row=2, values_only=True):
        carrera, cupos = fila[0], fila[1]

        carreras.append(
            Carrera(
                nombre=str(carrera).strip().upper(),
                cupos=int(cupos),
                puntaje_minimo=PUNTAJE_MINIMO_GENERAL
            )
        )

    return carreras


# ===============================
# ASIGNACIÓN SENESCYT
# ===============================
class AsignacionSenescyt:
    def asignar(self, aspirantes, carreras, historial):
        print(f"\nAspirantes cargados: {len(aspirantes)}")
        print(f"Carreras cargadas: {len(carreras)}")

        rondas = max(len(a.opciones) for a in aspirantes)

        for ronda in range(rondas):
            print(f"\n===== RONDA {ronda + 1} =====")

            for carrera in carreras:
                if carrera.cupos <= 0:
                    continue

                postulantes = [
                    a for a in aspirantes
                    if a.estado == "PENDIENTE"
                    and ronda < len(a.opciones)
                    and a.opciones[ronda] == carrera.nombre
                    and a.puntaje_final() >= carrera.puntaje_minimo
                ]

                postulantes.sort(
                    key=lambda a: a.puntaje_final(),
                    reverse=True
                )

                for aspirante in postulantes:
                    if carrera.cupos <= 0:
                        break

                    carrera.cupos -= 1
                    aspirante.carrera = carrera.nombre
                    aspirante.estado = "ASIGNADO"

                    historial.append(
                        (aspirante.nombre, carrera.nombre, ronda + 1)
                    )

                    print(
                        f"{aspirante.nombre} → {carrera.nombre} | "
                        f"Final: {aspirante.puntaje_final()}"
                    )


# ===============================
# SISTEMA (SINGLETON)
# ===============================
class Sistema:
    _instance = None

    def __init__(self):
        self.aspirantes = cargar_aspirantes_excel(
            "EJEMPLO_MATRIZ_ASIGNACIÓN_2025-2 (1).xlsx"
        )

        self.carreras = cargar_carreras_excel(
            "102_ULEAM_OFERTASEGMENTADA_IIPA-2025 (2).xlsx"
        )

        self.historial = []
        self.strategy = AsignacionSenescyt()

    @staticmethod
    def get_instance():
        if Sistema._instance is None:
            Sistema._instance = Sistema()
        return Sistema._instance

    def ejecutar(self):
        self.strategy.asignar(self.aspirantes, self.carreras, self.historial)
        print("\n✔ PROCESO FINALIZADO")

    def mostrar_historial(self):
        print("\n--- HISTORIAL ---")
        for i, h in enumerate(self.historial, 1):
            print(f"{i}. {h[0]} → {h[1]} (opción {h[2]})")


# ===============================
# MAIN
# ===============================
if __name__ == "__main__":
    sistema = Sistema.get_instance()
    sistema.ejecutar()
    sistema.mostrar_historial()